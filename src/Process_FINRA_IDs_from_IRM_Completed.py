"""
The purpose of this script is to run the FINRA IDs of a select group of individuals on the IRM
"""

import openpyxl
from datetime import datetime
import win32com.client as win32
import traceback
import time
from Generate_Header_Dictionary import get_column_headers
from FINRA_Scrape import search_webpage
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

start_time = time.time()

try:
    excel_app = win32.Dispatch("Excel.Application")
    excel_app.Visible = True
    print("Excel Application Initialized Successfully")
except Exception as e:
    print(f"Error initializing Excel Application: {e}")


#load the workbook and sheet
workbook_path = r"K:\Market Maps\Interest Rates Map.xlsm"
output_workbook_path = r"C:\Users\Bay Street - Larry B\Documents\Brielle\Programming\Projects\WebScrapers\workbooks\FINRA_Check_Output_Book.xlsm"
sheet_name = "Master"
table_name = "Master"
wb = openpyxl.load_workbook(workbook_path, data_only=True)
sheet = wb[sheet_name]

#group_search_value = "Interest Rate Swaps" #Add this line back in if you want to filter by group


def process_row(row, column_headers):
    finra_id = str(row[column_headers['FINRA ID']])
    
    #Check for specified conditions
    if finra_id.isdigit():
        output = search_webpage(finra_id)
        equivalent = False

        # Correct syntax differences
        if row[column_headers['Firm']][:4].lower() == output[:4].lower():
            equivalent = True
        elif row[column_headers['Firm']] == "Bank of America" and output == "BOFA SECURITIES, INC.":
            equivalent = True
        elif row[column_headers['Firm']][:7] == "Societe" and output == "SG AMERICAS SECURITIES, LLC":
            equivalent = True
        elif row[column_headers['Firm']][:3] == "JPM" and output[:4] == "J.P.":
            equivalent = True
        elif row[column_headers['Firm']] == "Pending" and output == "Inactive":
            equivalent = True

        # Check if firm is the same as on file
        if row[column_headers['Firm']].lower() != output.lower() and equivalent != True:
            time_of_check = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Create a dictionary for the row
            return {
                "Map Firm": row[column_headers['Firm']],
                "Name": row[column_headers['Name']],
                "Title": row[column_headers['Title']],
                "Function": row[column_headers['Function']],
                "Group": row[column_headers['Group']],
                "FINRA ID": finra_id,
                "Output": output,
                "Time of Check": time_of_check
            }
    return None


# Get column headers
column_headers = get_column_headers(workbook_path, sheet_name, table_name)

# Initialize an empty list to store the results
results = []


# Process rows using ThreadPoolExecutor
with ThreadPoolExecutor(max_workers=8) as executor:
    futures = [executor.submit(process_row, row, column_headers) for row in sheet.iter_rows(min_row=2, values_only=True)]
    for future in as_completed(futures):
        result = future.result()
        if result:
            results.append(result)

#Create a new workbook for output
output_wb = openpyxl.Workbook()
output_ws = output_wb.active

# Write headers
headers = ["Map Firm", "Name", "Title", "Function", "Group", "FINRA ID", "Output", "Time of Check"]
output_ws.append(headers)

# Write data rows
for result in results:
    row = [result[header] for header in headers]
    output_ws.append(row)

# Save the output workbook
output_wb.save(output_workbook_path)
output_wb.close()
wb.close()

#Path to the macro book
Macro_wb_path = r"C:\Users\Bay Street - Larry B\Documents\Brielle\Programming\Projects\WebScrapers\workbooks\Copy of FINRA_IRM_Report_Macro_Book.xlsm"

# Macro names
AFR_Macro = 'Attach_FINRA_Report.CreateEmailFromData'
Refresh_Macro = 'RefreshConnections.RefreshAllConnections'

try:
    # Start an instance of Excel
    excel_app = win32.Dispatch('Excel.Application')
    excel_app.Visible = False  # Or set to True if you want to see what happens in Excel

    # Open the workbook
    macro_wb = excel_app.Workbooks.Open(Filename=Macro_wb_path)

    # Run the macros
    excel_app.Run(Refresh_Macro)
    time.sleep(15)
    excel_app.Run(AFR_Macro)

    # Save the workbook
    macro_wb.Save()

# If there was an error, print the traceback
except Exception as e:
    print("An error occurred:")
    traceback.print_exc()

finally:
    # Close the workbook without saving again
    macro_wb.Close(False)

    # Quit Excel
    excel_app.Quit()

    # Cleanup the COM object
    del excel_app

end_time = time.time()
runtime = end_time - start_time
print(f"Script completed in {runtime:.2f} seconds")